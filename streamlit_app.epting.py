import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
import re

def load_excel(file_path):
    """Loads an Excel document into a pandas DataFrame."""
    return pd.read_excel(file_path, header=None)

def format_date(value):
    """Formats a date value to MM/DD/YYYY."""
    try:
        if isinstance(value, str):
            date_obj = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        elif isinstance(value, datetime):
            date_obj = value
        else:
            return value
        return date_obj.strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return value

def extract_value(row, key_index, key_name, all_keys):
    """Extracts the value associated with a key based on specific rules."""
    if key_name in ["AC", "FAU", "COIL", "HP", "AH"]:
        target_index = key_index + 1
        if target_index < len(row) and pd.notna(row[target_index]):
            return str(row[target_index]).strip()
    elif key_name == "LOT":
        for offset in range(1, len(row) - key_index):
            target_index = key_index + offset
            if target_index < len(row) and pd.notna(row[target_index]):
                value = row[target_index]
                if re.match(r"^\d+\s+.+", str(value), re.IGNORECASE):
                    return str(value).strip()
    elif key_name == "DATE":
        for offset in range(1, len(row) - key_index):
            target_index = key_index + offset
            if target_index < len(row) and pd.notna(row[target_index]):
                return format_date(row[target_index])
    elif key_name == "BUILDER":
        for offset in range(1, len(row) - key_index):
            target_index = key_index + offset
            if target_index < len(row) and pd.notna(row[target_index]):
                value = row[target_index]
                if re.match(r"^[A-Za-z\s]+$", str(value)):
                    return str(value).strip()
    return ""

def save_to_excel_with_left_quantities(output_path, layout, dataframe, all_keys):
    """Saves the provided data to a styled Excel document."""
    wb = Workbook()
    ws = wb.active
    header_font = Font(size=14, bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center")

    for key, (row, col) in layout.items():
        associated_data = ""
        for index, df_row in dataframe.iterrows():
            for col_index, value in enumerate(df_row):
                transformed_key = str(value).strip().upper()
                if transformed_key.lower() == key.strip().lower():
                    associated_data = extract_value(df_row, col_index, key, all_keys)
                    break
        ws.cell(row=row, column=col, value=key).font = header_font
        ws.cell(row=row, column=col + 1, value=associated_data).alignment = alignment

    wb.save(output_path)

# Streamlit Application
st.title("Excel Processor App")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
if uploaded_file:
    st.success("File uploaded successfully!")
    data = load_excel(uploaded_file)
    st.write("Preview of Uploaded Data:", data.head())

    layout = {
        "LOT": (1, 1),
        "DATE": (2, 1),
        "BUILDER": (3, 1),
        "TECH": (4, 1),
        "AC": (6, 1),
        "FAU": (7, 1),
        "COIL": (8, 1),
        "HP": (9, 1),
        "AH": (10, 1),
    }

    output_path = st.text_input("Enter a path to save the processed Excel file:")
    if st.button("Process Excel File"):
        all_keys = set(layout.keys())
        save_to_excel_with_left_quantities(output_path, layout, data, all_keys)
        st.success(f"Processing complete! File saved to {output_path}")
